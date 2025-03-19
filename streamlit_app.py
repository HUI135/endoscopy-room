import pandas as pd
import openpyxl
import streamlit as st
from io import BytesIO
from collections import defaultdict
import random
import time
import os
from datetime import datetime
from collections import Counter
import numpy as np
import re
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment

# ê³ ìœ í•œ ì‹œë“œ ìƒì„±
random.seed(time.time_ns() ^ int.from_bytes(os.urandom(4), 'big'))

st.header("ë‚´ì‹œê²½ ìŠ¤ì¼€ì¥´ ë°©ë°°ì • ë„êµ¬", divider='rainbow')
st.write(" ")

# ìµœëŒ€ í•œê³„ê°’ ì…ë ¥ UI
st.sidebar.header("ìµœëŒ€ ë°°ì • í•œê³„ ì„¤ì •")
MAX_DUTY = st.sidebar.number_input("1. ìµœëŒ€ ë‹¹ì§ í•©ê³„", min_value=1, value=3, step=1)
MAX_EARLY = st.sidebar.number_input("2. ìµœëŒ€ ì´ë¥¸ë°© í•©ê³„", min_value=1, value=6, step=1)
MAX_LATE = st.sidebar.number_input("3. ìµœëŒ€ ëŠ¦ì€ë°© í•©ê³„", min_value=1, value=6, step=1)
MAX_ROOM = st.sidebar.number_input("4. ìµœëŒ€ ë°©ë³„ í•©ê³„", min_value=1, value=3, step=1)

uploaded_file = st.file_uploader("ì—‘ì…€ íŒŒì¼ì„ ì—…ë¡œë“œí•˜ì„¸ìš” (Sheet1ê³¼ Sheet2 í¬í•¨)", type=["xlsx"])
if uploaded_file is not None:
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    Sheet1 = wb['Sheet1']
    Sheet2 = wb['Sheet2']

    def extract_data(sheet):
        data = {}
        headers = [cell.value for cell in sheet[1]]
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), 2):
            date_cell = row[0]
            row_values = [cell.value for cell in row]
            
            if date_cell.value:
                if isinstance(date_cell.value, datetime):
                    date = date_cell.value.date()
                    date_str = date.strftime('%Y-%m-%d')
                else:
                    date_str_raw = str(date_cell.value).strip()
                    try:
                        if "ì›”" in date_str_raw and "ì¼" in date_str_raw:
                            month, day = date_str_raw.replace("ì›”", "").replace("ì¼", "").split()
                            year = datetime.today().year
                            date = datetime.strptime(f"{year}-{month}-{day}", "%Y-%m-%d").date()
                            date_str = date.strftime('%Y-%m-%d')
                        else:
                            date = datetime.strptime(date_str_raw, '%Y-%m-%d').date()
                            date_str = date.strftime('%Y-%m-%d')
                    except ValueError:
                        continue
                
                # ì¤‘ë³µ ë‚ ì§œ ì²˜ë¦¬: ë™ì¼í•œ ë‚ ì§œê°€ ìˆìœ¼ë©´ ìŠ¤í‚µ
                if date_str in data:
                    continue
                
                day_of_week_raw = row[1].value if row[1].value else ""
                weekday_map = {
                    'ì›”': 'ì›”ìš”ì¼', 'í™”': 'í™”ìš”ì¼', 'ìˆ˜': 'ìˆ˜ìš”ì¼', 'ëª©': 'ëª©ìš”ì¼', 
                    'ê¸ˆ': 'ê¸ˆìš”ì¼', 'í† ': 'í† ìš”ì¼', 'ì¼': 'ì¼ìš”ì¼',
                    'Mon': 'ì›”ìš”ì¼', 'Tue': 'í™”ìš”ì¼', 'Wed': 'ìˆ˜ìš”ì¼', 'Thu': 'ëª©ìš”ì¼', 
                    'Fri': 'ê¸ˆìš”ì¼', 'Sat': 'í† ìš”ì¼', 'Sun': 'ì¼ìš”ì¼',
                    'Monday': 'ì›”ìš”ì¼', 'Tuesday': 'í™”ìš”ì¼', 'Wednesday': 'ìˆ˜ìš”ì¼', 
                    'Thursday': 'ëª©ìš”ì¼', 'Friday': 'ê¸ˆìš”ì¼', 'Saturday': 'í† ìš”ì¼', 'Sunday': 'ì¼ìš”ì¼'
                }
                day_of_week = day_of_week_raw
                for key, value in weekday_map.items():
                    if key in str(day_of_week_raw):
                        day_of_week = value
                        break
                else:
                    weekday_num = date.weekday()
                    weekdays = ['ì›”ìš”ì¼', 'í™”ìš”ì¼', 'ìˆ˜ìš”ì¼', 'ëª©ìš”ì¼', 'ê¸ˆìš”ì¼', 'í† ìš”ì¼', 'ì¼ìš”ì¼']
                    day_of_week = weekdays[weekday_num]
                
                personnel = []
                memo_dict = {}
                for cell in row[2:]:
                    if cell.value and cell.value not in ['ì›”ìš”ì¼', 'í™”ìš”ì¼', 'ìˆ˜ìš”ì¼', 'ëª©ìš”ì¼', 'ê¸ˆìš”ì¼', 'í† ìš”ì¼', 'ì¼ìš”ì¼']:
                        personnel.append(cell.value)
                    if cell.comment and cell.value:
                        memo_dict[cell.value] = cell.comment.text.strip()

                personnel_with_suffix = []
                name_counts = Counter()
                for name in personnel:
                    name_counts[name] += 1
                    suffix = f"_{name_counts[name]}" if name_counts[name] > 1 else ""
                    personnel_with_suffix.append(f"{name}{suffix}")
                
                data[date_str] = {
                    'personnel': personnel_with_suffix, 
                    'original_personnel': personnel, 
                    'day': day_of_week, 
                    'memos': memo_dict, 
                    'headers': headers
                }
        
        return data

    Sheet1_data = extract_data(Sheet1)
    Sheet2_data = extract_data(Sheet2)

    if not Sheet1_data:
        st.error("Sheet1_dataê°€ ë¹„ì–´ ìˆìŠµë‹ˆë‹¤. ì—‘ì…€ íŒŒì¼ì˜ Sheet1ì— ë°ì´í„°ê°€ ìˆëŠ”ì§€, í˜•ì‹ì´ ë§ëŠ”ì§€ í™•ì¸í•˜ì„¸ìš”.")
        st.stop()

    # Sheet1ê³¼ Sheet2 ê°„ì˜ ì¸ì› ë¶ˆì¼ì¹˜ ê°ì§€
    mismatch_warnings = []
    for date in Sheet2_data.keys():
        if date in Sheet1_data:
            sheet1_personnel = set(Sheet1_data[date]['original_personnel'])
            fixed_assignments = {}
            for row in Sheet2.iter_rows(min_row=2):
                sheet2_date = row[0].value
                if sheet2_date:
                    if isinstance(sheet2_date, datetime):
                        date_str = sheet2_date.strftime('%Y-%m-%d')
                    else:
                        date_str_raw = str(sheet2_date).strip()
                        try:
                            if "ì›”" in date_str_raw and "ì¼" in date_str_raw:
                                month, day = date_str_raw.replace("ì›”", "").replace("ì¼", "").split()
                                year = datetime.today().year
                                date = datetime.strptime(f"{year}-{month}-{day}", "%Y-%m-%d").date()
                                date_str = date.strftime('%Y-%m-%d')
                            else:
                                date = datetime.strptime(date_str_raw, '%Y-%m-%d').date()
                                date_str = date.strftime('%Y-%m-%d')
                        except ValueError:
                            continue
                    if date_str == date:
                        headers = Sheet2_data[date_str]['headers']
                        for col_idx, cell in enumerate(row[2:], 2):
                            if cell.value:
                                slot_key = headers[col_idx]
                                fixed_assignments[cell.value] = slot_key

            # Sheet2ì— ê³ ì • ë°°ì¹˜ëœ ì¸ì›ì´ Sheet1ì— ì—†ëŠ”ì§€ í™•ì¸
            for person, slot in fixed_assignments.items():
                if person not in sheet1_personnel:
                    date_obj = datetime.strptime(date, '%Y-%m-%d')
                    formatted_date = date_obj.strftime('%mì›” %dì¼')
                    mismatch_warnings.append(
                        f"Sheet1ì˜ {formatted_date}ì—ëŠ” '{person}'ì´ ì—†ìŒì—ë„, Sheet2ì˜ {formatted_date}ì— '{person}'ì´ '{slot}'ì— ë°°ì¹˜ë˜ì–´ ìˆìŠµë‹ˆë‹¤. "
                        f"ì´ ê²½ìš° {formatted_date}ì˜ Sheet1 ì¸ì›ì´ ì™„ì „íˆ ë°°ì¹˜ë˜ì§€ ì•Šì„ ìˆ˜ ìˆìŠµë‹ˆë‹¤."
                    )

    # ê²½ê³  ë©”ì‹œì§€ ì¶œë ¥
    if mismatch_warnings:
        for warning in mismatch_warnings:
            st.warning(warning)

    def apply_memo_rules(assignment, personnel, memos, fixed_personnel, slots, assigned_counts, personnel_counts, time_groups, assigned_by_time, total_early, total_late, total_duty, total_rooms, day_of_week, ignore_memos=None):
        if ignore_memos is None:
            ignore_memos = set()
        prioritized = []
        all_slots = set(slots)
        for person in personnel:
            original_name = person.split('_')[0]
            if original_name in memos and person not in fixed_personnel and original_name not in ignore_memos:
                rule = memos[original_name]
                if rule in memo_rules:
                    if rule in ['ë‹¹ì§ ì•ˆë¨', 'ì˜¤ì „ ë‹¹ì§ ì•ˆë¨', 'ì˜¤í›„ ë‹¹ì§ ì•ˆë¨']:
                        forbidden_slots = memo_rules[rule]
                        allowed_slots = list(all_slots - set(forbidden_slots))
                        prioritized.append((person, allowed_slots))
                    else:
                        prioritized.append((person, memo_rules[rule]))
        remaining_slots = [i for i, x in enumerate(assignment) if x is None]
        memo_assignments = {}
        for person, allowed_slots in prioritized:
            original_name = person.split('_')[0]
            valid_slots = [
                i for i in remaining_slots 
                if slots[i] in allowed_slots 
                and assigned_counts[person] < personnel_counts[person]
                and person not in assigned_by_time.get(next(t for t, g in time_groups.items() if slots[i] in g), set())
                and total_early[original_name] < MAX_EARLY
                and total_late[original_name] < MAX_LATE
                and total_duty[original_name] < MAX_DUTY
            ]
            if valid_slots:
                slot_idx = random.choice(valid_slots)
                assignment[slot_idx] = person
                assigned_counts[person] += 1
                memo_assignments.setdefault(slots[slot_idx], Counter())[person] += 1
                remaining_slots.remove(slot_idx)
                for time_group, group in time_groups.items():
                    if slots[slot_idx] in group:
                        assigned_by_time[time_group].add(person)
                if slots[slot_idx] in {'8:30(1)_ë‹¹ì§', '8:30(2)', '8:30(4)', '8:30(7)'} and day_of_week != 'í† ìš”ì¼':
                    total_early[original_name] += 1
                if slots[slot_idx] in {'10:00(9)', '10:00(3)'} and day_of_week != 'í† ìš”ì¼':
                    total_late[original_name] += 1
                if slots[slot_idx] in {'8:30(1)_ë‹¹ì§', '13:30(3)_ë‹¹ì§'} and day_of_week != 'í† ìš”ì¼':
                    total_duty[original_name] += 1
                room_num = re.search(r'\((\d+)\)', slots[slot_idx])
                if room_num and day_of_week != 'í† ìš”ì¼':
                    total_rooms[room_num.group(1)][original_name] += 1
        return assignment, memo_assignments

    def calculate_stats(assignment, slots, day_of_week):
        early_slots = {'8:30(1)_ë‹¹ì§', '8:30(2)', '8:30(4)', '8:30(7)'}
        late_slots = {'10:00(9)', '10:00(3)'}
        duty_slots = {'8:30(1)_ë‹¹ì§', '13:30(3)_ë‹¹ì§'}
        slot_counts = {slot.replace('_ë‹¹ì§', ''): Counter() for slot in time_slots.keys() if slot != 'ì˜¨ì½œ'}
        
        stats = Counter()
        early_count = Counter()
        late_count = Counter()
        duty_count = Counter()
        
        for slot, person in zip(slots, assignment):
            if person:
                original_name = person.split('_')[0]
                stats[original_name] += 1
                if day_of_week != 'í† ìš”ì¼':
                    if slot in early_slots:
                        early_count[original_name] += 1
                    if slot in late_slots:
                        late_count[original_name] += 1
                    if slot in duty_slots:
                        duty_count[original_name] += 1
                    if slot != 'ì˜¨ì½œ':
                        slot_counts[slot.replace('_ë‹¹ì§', '')][original_name] += 1
        
        return stats, early_count, late_count, duty_count, slot_counts

    def count_violations(total_early, total_late, total_duty, total_slots):
        violations = 0
        all_personnel = set(total_early.keys()) | set(total_late.keys()) | set(total_duty.keys()) | set().union(*[total_slots[slot].keys() for slot in total_slots])
        for person in all_personnel:
            if total_early.get(person, 0) > MAX_EARLY:
                violations += total_early.get(person, 0) - MAX_EARLY
            if total_late.get(person, 0) > MAX_LATE:
                violations += total_late.get(person, 0) - MAX_LATE
            if total_duty.get(person, 0) > MAX_DUTY:
                violations += total_duty.get(person, 0) - MAX_DUTY
            for slot in total_slots:
                if total_slots[slot].get(person, 0) > MAX_ROOM:
                    violations += total_slots[slot].get(person, 0) - MAX_ROOM
        return violations

    def random_assign(personnel, slots, fixed_assignments, memos, day_of_week, time_groups, total_stats, current_date):
        random.seed(time.time_ns() ^ int.from_bytes(os.urandom(4), 'big'))
        
        max_attempts = 100
        duty_slots = ['8:30(1)_ë‹¹ì§', '13:30(3)_ë‹¹ì§']
        early_slots = ['8:30(1)_ë‹¹ì§', '8:30(2)', '8:30(4)', '8:30(7)']
        late_slots = ['10:00(9)', '10:00(3)']
        
        best_assignment = None
        best_fixed_assignments_record = None
        best_memo_assignments = None
        min_violations = float('inf')
        best_total_early = total_stats['early'].copy()
        best_total_late = total_stats['late'].copy()
        best_total_duty = total_stats['duty'].copy()
        best_total_slots = {slot: total_stats['slots'][slot].copy() for slot in total_stats['slots']}
        best_total_stats = total_stats['total'].copy()

        for attempt in range(max_attempts):
            assignment = [None] * len(slots)
            fixed_personnel = set()
            assigned_counts = Counter()
            personnel_counts = Counter(personnel)
            assigned_by_time = {time_group: set() for time_group in time_groups.keys()}
            fixed_assignments_record = {}
            memo_assignments = {}
            
            total_early = total_stats['early'].copy()
            total_late = total_stats['late'].copy()
            total_duty = total_stats['duty'].copy()
            total_rooms = {str(i): total_stats['rooms'][str(i)].copy() for i in range(1, 13)}
            
            # ê³ ì • ë°°ì¹˜ ì ìš©
            for date, assignments in fixed_assignments.items():
                if date == current_date:
                    for person, fixed_slot in assignments.items():
                        if fixed_slot in slots:
                            slot_idx = slots.index(fixed_slot)
                            original_name = person.split('_')[0]
                            time_group = next(t for t, g in time_groups.items() if fixed_slot in g)
                            if person in assigned_by_time[time_group]:
                                continue  # ì¤‘ë³µ ë°°ì • ë°©ì§€
                            assignment[slot_idx] = person
                            fixed_personnel.add(person)
                            assigned_counts[person] += 1
                            fixed_assignments_record.setdefault(fixed_slot, Counter())[person] += 1
                            assigned_by_time[time_group].add(person)
                            if fixed_slot in early_slots and day_of_week != 'í† ìš”ì¼':
                                total_early[original_name] += 1
                            if fixed_slot in late_slots and day_of_week != 'í† ìš”ì¼':
                                total_late[original_name] += 1
                            if fixed_slot in duty_slots and day_of_week != 'í† ìš”ì¼':
                                total_duty[original_name] += 1
                            room_num = re.search(r'\((\d+)\)', fixed_slot)
                            if room_num and day_of_week != 'í† ìš”ì¼':
                                total_rooms[room_num.group(1)][original_name] += 1

            # ë©”ëª¨ ê¸°ë°˜ ìš°ì„  ë°°ì¹˜
            all_slots = set(slots)
            prioritized = []
            for person in personnel:
                original_name = person.split('_')[0]
                if original_name in memos and person not in fixed_personnel:
                    rule = memos[original_name]
                    if rule in memo_rules:
                        if rule in ['ë‹¹ì§ ì•ˆë¨', 'ì˜¤ì „ ë‹¹ì§ ì•ˆë¨', 'ì˜¤í›„ ë‹¹ì§ ì•ˆë¨']:
                            forbidden_slots = memo_rules[rule]
                            allowed_slots = list(all_slots - set(forbidden_slots))
                            prioritized.append((person, allowed_slots))
                        else:
                            prioritized.append((person, memo_rules[rule]))
            
            remaining_slots = [i for i, x in enumerate(assignment) if x is None]
            for person, allowed_slots in prioritized:
                original_name = person.split('_')[0]
                valid_slots = [
                    i for i in remaining_slots 
                    if slots[i] in allowed_slots 
                    and assigned_counts[person] < personnel_counts[person]
                    and person not in assigned_by_time.get(next(t for t, g in time_groups.items() if slots[i] in g), set())
                    and total_early[original_name] < MAX_EARLY
                    and total_late[original_name] < MAX_LATE
                    and total_duty[original_name] < MAX_DUTY
                ]
                if valid_slots:
                    slot_idx = random.choice(valid_slots)
                    assignment[slot_idx] = person
                    assigned_counts[person] += 1
                    memo_assignments.setdefault(slots[slot_idx], Counter())[person] += 1
                    remaining_slots.remove(slot_idx)
                    for time_group, group in time_groups.items():
                        if slots[slot_idx] in group:
                            assigned_by_time[time_group].add(person)
                    if slots[slot_idx] in early_slots and day_of_week != 'í† ìš”ì¼':
                        total_early[original_name] += 1
                    if slots[slot_idx] in late_slots and day_of_week != 'í† ìš”ì¼':
                        total_late[original_name] += 1
                    if slots[slot_idx] in duty_slots and day_of_week != 'í† ìš”ì¼':
                        total_duty[original_name] += 1
                    room_num = re.search(r'\((\d+)\)', slots[slot_idx])
                    if room_num and day_of_week != 'í† ìš”ì¼':
                        total_rooms[room_num.group(1)][original_name] += 1

            # ë‹¹ì§ ìŠ¬ë¡¯ ë°°ì •
            available_slots = [i for i, slot in enumerate(slots) if assignment[i] is None]
            personnel_list = [p for p in personnel if assigned_counts[p] < personnel_counts[p]]
            duty_indices = [i for i in available_slots if slots[i] in duty_slots]
            personnel_list = sorted(personnel_list, key=lambda p: total_duty[p.split('_')[0]])
            for slot_idx in duty_indices:
                time_group = next(t for t, g in time_groups.items() if slots[slot_idx] in g)
                for person in personnel_list:
                    original_name = person.split('_')[0]
                    if (person not in assigned_by_time[time_group] and 
                        assigned_counts[person] < personnel_counts[person] and
                        total_duty[original_name] < MAX_DUTY):
                        assignment[slot_idx] = person
                        assigned_counts[person] += 1
                        assigned_by_time[time_group].add(person)
                        if day_of_week != 'í† ìš”ì¼':
                            total_duty[original_name] += 1
                            total_early[original_name] += 1
                        room_num = re.search(r'\((\d+)\)', slots[slot_idx])
                        if room_num and day_of_week != 'í† ìš”ì¼':
                            total_rooms[room_num.group(1)][original_name] += 1
                        available_slots.remove(slot_idx)
                        personnel_list = [p for p in personnel_list if assigned_counts[p] < personnel_counts[p]]
                        break

            # ì´ë¥¸ë°© ìŠ¬ë¡¯ ë°°ì •
            early_indices = [i for i in available_slots if slots[i] in early_slots]
            personnel_list = sorted(personnel_list, key=lambda p: total_early[p.split('_')[0]])
            for slot_idx in early_indices:
                time_group = next(t for t, g in time_groups.items() if slots[slot_idx] in g)
                for person in personnel_list:
                    original_name = person.split('_')[0]
                    if (person not in assigned_by_time[time_group] and 
                        assigned_counts[person] < personnel_counts[person] and
                        total_early[original_name] < MAX_EARLY):
                        assignment[slot_idx] = person
                        assigned_counts[person] += 1
                        assigned_by_time[time_group].add(person)
                        if day_of_week != 'í† ìš”ì¼':
                            total_early[original_name] += 1
                        room_num = re.search(r'\((\d+)\)', slots[slot_idx])
                        if room_num and day_of_week != 'í† ìš”ì¼':
                            total_rooms[room_num.group(1)][original_name] += 1
                        available_slots.remove(slot_idx)
                        personnel_list = [p for p in personnel_list if assigned_counts[p] < personnel_counts[p]]
                        break

            # ëŠ¦ì€ë°© ìŠ¬ë¡¯ ë°°ì •
            late_indices = [i for i in available_slots if slots[i] in late_slots]
            personnel_list = sorted(personnel_list, key=lambda p: total_late[p.split('_')[0]])
            for slot_idx in late_indices:
                time_group = next(t for t, g in time_groups.items() if slots[slot_idx] in g)
                for person in personnel_list:
                    original_name = person.split('_')[0]
                    if (person not in assigned_by_time[time_group] and 
                        assigned_counts[person] < personnel_counts[person] and
                        total_late[original_name] < MAX_LATE):
                        assignment[slot_idx] = person
                        assigned_counts[person] += 1
                        assigned_by_time[time_group].add(person)
                        if day_of_week != 'í† ìš”ì¼':
                            total_late[original_name] += 1
                        room_num = re.search(r'\((\d+)\)', slots[slot_idx])
                        if room_num and day_of_week != 'í† ìš”ì¼':
                            total_rooms[room_num.group(1)][original_name] += 1
                        available_slots.remove(slot_idx)
                        personnel_list = [p for p in personnel_list if assigned_counts[p] < personnel_counts[p]]
                        break

            # ë‚˜ë¨¸ì§€ ìŠ¬ë¡¯ ë°°ì •
            available_slots = [i for i, slot in enumerate(slots) if assignment[i] is None]
            personnel_list = [p for p in personnel if assigned_counts[p] < personnel_counts[p]]
            random.shuffle(personnel_list)
            assignment, available_slots = assign_remaining(assignment, personnel_list, available_slots, slots, assigned_counts, personnel_counts, time_groups, assigned_by_time, total_early, total_late, total_duty, total_rooms, MAX_EARLY, MAX_LATE, MAX_DUTY, MAX_ROOM, day_of_week)

            # ê°•ì œ ë°°ì •
            if available_slots:
                personnel_list = sorted(
                    personnel_list,
                    key=lambda p: (total_duty[p.split('_')[0]], total_early[p.split('_')[0]], total_late[p.split('_')[0]], sum(total_rooms[r][p.split('_')[0]] for r in total_rooms))
                )
                for slot_idx in available_slots:
                    time_group = next(t for t, g in time_groups.items() if slots[slot_idx] in g)
                    for person in personnel_list:
                        original_name = person.split('_')[0]
                        if (assigned_counts[person] < personnel_counts[person] and 
                            person not in assigned_by_time[time_group]):
                            assignment[slot_idx] = person
                            assigned_counts[person] += 1
                            assigned_by_time[time_group].add(person)
                            if slots[slot_idx] in early_slots and day_of_week != 'í† ìš”ì¼':
                                total_early[original_name] += 1
                            if slots[slot_idx] in late_slots and day_of_week != 'í† ìš”ì¼':
                                total_late[original_name] += 1
                            if slots[slot_idx] in duty_slots and day_of_week != 'í† ìš”ì¼':
                                total_duty[original_name] += 1
                            room_num = re.search(r'\((\d+)\)', slots[slot_idx])
                            if room_num and day_of_week != 'í† ìš”ì¼':
                                total_rooms[room_num.group(1)][original_name] += 1
                            available_slots.remove(slot_idx)
                            break

            # í†µê³„ ê³„ì‚° ë° ìœ„ë°˜ í™•ì¸
            stats, early_count, late_count, duty_count, slot_counts = calculate_stats(assignment, slots, day_of_week)
            temp_total_early = total_stats['early'].copy()
            temp_total_late = total_stats['late'].copy()
            temp_total_duty = total_stats['duty'].copy()
            temp_total_slots = {slot: total_stats['slots'][slot].copy() for slot in total_stats['slots']}
            temp_total_stats = total_stats['total'].copy()

            temp_total_early.update(early_count)
            temp_total_late.update(late_count)
            temp_total_duty.update(duty_count)
            for slot in slot_counts:
                temp_total_slots[slot].update(slot_counts[slot])
            temp_total_stats.update(stats)

            violations = count_violations(temp_total_early, temp_total_late, temp_total_duty, temp_total_slots)

            if violations < min_violations:
                min_violations = violations
                best_assignment = assignment.copy()
                best_fixed_assignments_record = fixed_assignments_record.copy()
                best_memo_assignments = memo_assignments.copy()
                best_total_early = temp_total_early.copy()
                best_total_late = temp_total_late.copy()
                best_total_duty = temp_total_duty.copy()
                best_total_slots = {slot: temp_total_slots[slot].copy() for slot in temp_total_slots}
                best_total_stats = temp_total_stats.copy()
                if min_violations == 0:
                    break

        if best_assignment is not None:
            total_stats['early'] = best_total_early
            total_stats['late'] = best_total_late
            total_stats['duty'] = best_total_duty
            total_stats['slots'] = best_total_slots
            total_stats['total'] = best_total_stats
            return best_assignment, best_fixed_assignments_record, best_memo_assignments
        
        stats, early_count, late_count, duty_count, slot_counts = calculate_stats(assignment, slots, day_of_week)
        total_stats['early'].update(early_count)
        total_stats['late'].update(late_count)
        total_stats['duty'].update(duty_count)
        for slot in slot_counts:
            total_stats['slots'][slot].update(slot_counts[slot])
        total_stats['total'].update(stats)
        return assignment, fixed_assignments_record, memo_assignments

    def assign_remaining(assignment, personnel_list, available_slots, slots, assigned_counts, personnel_counts, time_groups, assigned_by_time, total_early, total_late, total_duty, total_rooms, MAX_EARLY, MAX_LATE, MAX_DUTY, MAX_ROOM, day_of_week):
        random.shuffle(personnel_list)
        for person in personnel_list:
            if available_slots:
                original_name = person.split('_')[0]
                possible_slots = []
                
                for slot_idx in available_slots:
                    slot = slots[slot_idx]
                    time_group = next(t for t, g in time_groups.items() if slot in g)
                    if person not in assigned_by_time[time_group]:
                        early_ok = (total_early[original_name] + (1 if slot in {'8:30(1)_ë‹¹ì§', '8:30(2)', '8:30(4)', '8:30(7)'} else 0)) <= MAX_EARLY or day_of_week == 'í† ìš”ì¼'
                        late_ok = (total_late[original_name] + (1 if slot in {'10:00(9)', '10:00(3)'} else 0)) <= MAX_LATE or day_of_week == 'í† ìš”ì¼'
                        duty_ok = (total_duty[original_name] + (1 if slot in {'8:30(1)_ë‹¹ì§', '13:30(3)_ë‹¹ì§'} else 0)) <= MAX_DUTY or day_of_week == 'í† ìš”ì¼'
                        room_num = re.search(r'\((\d+)\)', slot)
                        room_ok = True
                        if room_num:
                            room = room_num.group(1)
                            room_ok = (total_rooms[room][original_name] + 1) <= MAX_ROOM or day_of_week == 'í† ìš”ì¼'
                            if not room_ok and (day_of_week != 'í† ìš”ì¼' or time_group != '9:00'):
                                group = next(g for t, g in time_groups.items() if slot in g)
                                alt_slots = [s for s in group if s != slot and s in slots and slots.index(s) in available_slots]
                                for alt_slot in alt_slots:
                                    alt_idx = slots.index(alt_slot)
                                    alt_room_num = re.search(r'\((\d+)\)', alt_slot)
                                    if alt_room_num and (total_rooms[alt_room_num.group(1)][original_name] < MAX_ROOM or day_of_week == 'í† ìš”ì¼'):
                                        possible_slots.append(alt_idx)
                                        break
                                continue
                        
                        if early_ok and late_ok and duty_ok and room_ok:
                            possible_slots.append(slot_idx)
                
                if possible_slots:
                    slot_idx = random.choice(possible_slots)
                    assignment[slot_idx] = person
                    assigned_counts[person] += 1
                    for time_group, group in time_groups.items():
                        if slots[slot_idx] in group:
                            assigned_by_time[time_group].add(person)
                    if slots[slot_idx] in {'8:30(1)_ë‹¹ì§', '8:30(2)', '8:30(4)', '8:30(7)'} and day_of_week != 'í† ìš”ì¼':
                        total_early[original_name] += 1
                    if slots[slot_idx] in {'10:00(9)', '10:00(3)'} and day_of_week != 'í† ìš”ì¼':
                        total_late[original_name] += 1
                    if slots[slot_idx] in {'8:30(1)_ë‹¹ì§', '13:30(3)_ë‹¹ì§'} and day_of_week != 'í† ìš”ì¼':
                        total_duty[original_name] += 1
                    room_num = re.search(r'\((\d+)\)', slots[slot_idx])
                    if room_num and day_of_week != 'í† ìš”ì¼':
                        total_rooms[room_num.group(1)][original_name] += 1
                    available_slots.remove(slot_idx)
        return assignment, available_slots

    time_slots = {
        '8:30(1)_ë‹¹ì§': 0, '8:30(2)': 1, '8:30(4)': 2, '8:30(7)': 3,
        '9:00(10)': 4, '9:00(11)': 5, '9:00(12)': 6,
        '9:30(8)': 7, '9:30(5)': 8, '9:30(6)': 9,
        '10:00(9)': 10, '10:00(3)': 11,
        'ì˜¨ì½œ': 12,
        '13:30(3)_ë‹¹ì§': 13, '13:30(4)': 14, '13:30(9)': 15, '13:30(2)': 16
    }

    time_groups = {
        '8:30': ['8:30(1)_ë‹¹ì§', '8:30(2)', '8:30(4)', '8:30(7)'],
        '9:00': ['9:00(10)', '9:00(11)', '9:00(12)'],
        '9:30': ['9:30(8)', '9:30(5)', '9:30(6)'],
        '10:00': ['10:00(9)', '10:00(3)'],
        '13:30': ['13:30(3)_ë‹¹ì§', '13:30(4)', '13:30(9)', '13:30(2)'],
        'ì˜¨ì½œ': ['ì˜¨ì½œ']
    }

    weekday_slots = list(time_slots.keys())
    saturday_slots = ['8:30(1)_ë‹¹ì§', '8:30(2)', '8:30(4)', '8:30(7)', '9:00(10)', '9:30(8)', '9:30(5)', '9:30(6)', '10:00(9)', '10:00(3)']
    slot_mappings = {}
    for date, data in Sheet1_data.items():
        day_of_week = data['day']
        if day_of_week == 'í† ìš”ì¼':
            slot_mappings[date] = saturday_slots
        else:
            slot_mappings[date] = weekday_slots

    memo_rules = {
        'ë‹¹ì§ ì•ˆë¨': ['8:30(1)_ë‹¹ì§', '13:30(3)_ë‹¹ì§'],
        'ì˜¤ì „ ë‹¹ì§ ì•ˆë¨': ['8:30(1)_ë‹¹ì§'],
        'ì˜¤í›„ ë‹¹ì§ ì•ˆë¨': ['13:30(3)_ë‹¹ì§'],
        'ë‹¹ì§ ì•„ë‹Œ ì´ë¥¸ë°©': ['8:30(2)', '8:30(4)', '8:30(7)'],
        '8:30': ['8:30(2)', '8:30(4)', '8:30(7)'],
        '9:00': ['9:00(10)', '9:00(11)', '9:00(12)'],
        '9:30': ['9:30(8)', '9:30(5)', '9:30(6)'],
        '10:00': ['10:00(9)', '10:00(3)'],
        'ì´ë¥¸ë°©': ['8:30(1)_ë‹¹ì§', '8:30(2)', '8:30(4)', '8:30(7)'],
        'ì˜¤í›„ ë‹¹ì§': ['13:30(3)_ë‹¹ì§'],
        'ì˜¤ì „ ë‹¹ì§': ['8:30(1)_ë‹¹ì§']
    }

    # total_stats ì´ˆê¸°í™”
    if 'total_stats' not in st.session_state:
        st.session_state.total_stats = {
            'total': Counter(),
            'early': Counter(),
            'late': Counter(),
            'duty': Counter(),
            'slots': {slot.replace('_ë‹¹ì§', ''): Counter() for slot in time_slots.keys() if slot != 'ì˜¨ì½œ'},
            'rooms': {str(i): Counter() for i in range(1, 13)}
        }
    total_stats = st.session_state.total_stats

    # total_stats ì´ˆê¸°í™” (ì„¸ì…˜ ìƒíƒœê°€ ë³€ê²½ë  ë•Œë§ˆë‹¤ ì´ˆê¸°í™”)
    total_stats['total'].clear()
    total_stats['early'].clear()
    total_stats['late'].clear()
    total_stats['duty'].clear()
    for slot in total_stats['slots']:
        total_stats['slots'][slot].clear()
    for room in total_stats['rooms']:
        total_stats['rooms'][room].clear()

    # Sheet1_data ìˆœíšŒ ë° ë°°ì •
    assignments = {}
    fixed_assignments = {}
    for date in sorted(Sheet1_data.keys()):  # ë‚ ì§œ ì •ë ¬
        personnel = Sheet1_data[date]['personnel']
        day_of_week = Sheet1_data[date]['day']
        memos = Sheet1_data[date]['memos']
        
        fixed_assignments[date] = {}
        for person in personnel:
            original_name = person.split('_')[0]
            if date in Sheet2_data:
                for p, slot in Sheet2_data[date].get('fixed_assignments', {}).items():
                    if p == original_name and slot in time_slots:
                        fixed_assignments[date][person] = slot
        
        assigned_slots = slot_mappings.get(date, weekday_slots)
        assignment, fixed_assignments_record, memo_assignments = random_assign(
            personnel, assigned_slots, fixed_assignments, memos, day_of_week, time_groups, total_stats, current_date=date
        )
        assignments[date] = assignment

    # íŒŒì¼ ë³€ê²½ ê°ì§€ ë° ì„¸ì…˜ ì´ˆê¸°í™”
    file_hash = hash(uploaded_file.getvalue())
    if 'last_file_hash' not in st.session_state or st.session_state.last_file_hash != file_hash:
        st.session_state.clear()
        st.session_state.last_file_hash = file_hash

    # ì„¸ì…˜ ìƒíƒœ ì´ˆê¸°í™” ë° ë°°ì • ê³„ì‚°
    if 'assignments' not in st.session_state:
        assignments = {}
        slot_mappings = {}
        total_stats = {
            'total': Counter(), 
            'early': Counter(), 
            'late': Counter(), 
            'duty': Counter(), 
            'rooms': {str(i): Counter() for i in range(1, 13)},
            'slots': {slot.replace('_ë‹¹ì§', ''): Counter() for slot in time_slots.keys() if slot != 'ì˜¨ì½œ'}
        }
        total_fixed_stats = {slot: Counter() for slot in time_slots.keys()}
        total_memo_stats = {slot: Counter() for slot in time_slots.keys()}

        for date, data in sorted(Sheet1_data.items()):  # ë‚ ì§œ ì •ë ¬
            personnel = data['personnel']
            original_personnel = data['original_personnel']
            memos = data['memos']
            day_of_week = data['day']

            if day_of_week == 'í† ìš”ì¼':
                slots = saturday_slots.copy()
            else:
                slots = weekday_slots.copy()

            fixed_assignments = {}
            current_date = date
            for row in Sheet2.iter_rows(min_row=2):
                sheet2_date = row[0].value
                if sheet2_date:
                    if isinstance(sheet2_date, datetime):
                        date_str = sheet2_date.strftime('%Y-%m-%d')
                    else:
                        date_str_raw = str(sheet2_date).strip()
                        try:
                            if "ì›”" in date_str_raw and "ì¼" in date_str_raw:
                                month, day = date_str_raw.replace("ì›”", "").replace("ì¼", "").split()
                                year = datetime.today().year
                                date = datetime.strptime(f"{year}-{month}-{day}", "%Y-%m-%d").date()
                                date_str = date.strftime('%Y-%m-%d')
                            else:
                                date = datetime.strptime(date_str_raw, '%Y-%m-%d').date()
                                date_str = date.strftime('%Y-%m-%d')
                        except ValueError:
                            continue
                    fixed_assignments[date_str] = {}
                    if date_str in Sheet2_data:
                        headers = Sheet2_data[date_str]['headers']
                        for col_idx, cell in enumerate(row[2:], 2):
                            if cell.value:
                                slot_key = headers[col_idx]
                                # Sheet1ì— í•´ë‹¹ ì¸ì›ì´ ìˆëŠ” ê²½ìš°ì—ë§Œ ê³ ì • ë°°ì¹˜ ì¶”ê°€
                                if date_str in Sheet1_data and cell.value in Sheet1_data[date_str]['original_personnel']:
                                    fixed_assignments[date_str][cell.value] = slot_key

            if personnel:
                assignment, fixed_assignments_record, memo_assignments = random_assign(
                    personnel, slots, fixed_assignments, memos, day_of_week, time_groups, total_stats, current_date=date
                )
                assignments[date] = assignment
                slot_mappings[date] = slots
                
                for slot in fixed_assignments_record:
                    total_fixed_stats[slot].update(fixed_assignments_record[slot])
                for slot in memo_assignments:
                    total_memo_stats[slot].update(memo_assignments[slot])
            else:
                assignments[date] = [None] * len(slots)
                slot_mappings[date] = slots

        st.session_state.assignments = assignments
        st.session_state.slot_mappings = slot_mappings
        st.session_state.total_stats = total_stats
        st.session_state.total_fixed_stats = total_fixed_stats
        st.session_state.total_memo_stats = total_memo_stats
    else:
        assignments = st.session_state.assignments
        slot_mappings = st.session_state.slot_mappings
        total_stats = st.session_state.total_stats
        total_fixed_stats = st.session_state.total_fixed_stats
        total_memo_stats = st.session_state.total_memo_stats

    # í†µí•© ë°°ì¹˜ ê²°ê³¼ DataFrame
    result_data = []
    all_columns = ['ë‚ ì§œ', 'ìš”ì¼'] + list(time_slots.keys())
    memo_mapping = {}

    for date in sorted(Sheet1_data.keys()):  # ë‚ ì§œ ì •ë ¬
        assigned_slots = slot_mappings.get(date, weekday_slots)
        assignment = assignments.get(date, [None] * len(assigned_slots))
        memos = Sheet1_data[date]['memos']
        
        slot_to_person = {slot: None for slot in time_slots.keys()}
        memo_mapping[date] = {}

        # assigned_slotsì™€ assignment ë§¤í•‘
        for slot, person in zip(assigned_slots, assignment):
            if person:
                original_name = person.split('_')[0] if '_' in person else person
                slot_to_person[slot] = original_name
                if original_name in memos:
                    memo_mapping[date][(original_name, slot)] = memos[original_name]

        row = [date, Sheet1_data[date]['day']] + [slot_to_person[slot] for slot in time_slots.keys()]
        result_data.append(row)

    if not result_data:
        st.error("result_dataê°€ ë¹„ì–´ ìˆìŠµë‹ˆë‹¤. ë°°ì • ê²°ê³¼ê°€ ìƒì„±ë˜ì§€ ì•Šì•˜ìŠµë‹ˆë‹¤.")
        st.stop()

    result_df = pd.DataFrame(result_data, columns=all_columns)

    # ì¸ì›ë³„ ì „ì²´ í†µê³„ DataFrame
    all_personnel = set(total_stats['total'].keys())
    if not all_personnel:
        all_personnel = set().union(*[set(data['original_personnel']) for data in Sheet1_data.values()])
        if not all_personnel:
            st.error("ì¸ì› ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤. Sheet1_dataë¥¼ í™•ì¸í•˜ì„¸ìš”.")
            st.stop()

    stats_data = []
    slot_columns = [slot.replace('_ë‹¹ì§', '') for slot in time_slots.keys() if slot != 'ì˜¨ì½œ']
    for person in all_personnel:
        row = {
            'ì¸ì›': person,
            'ì „ì²´ í•©ê³„': total_stats['total'].get(person, 0),
            'ì´ë¥¸ë°© í•©ê³„': total_stats['early'].get(person, 0),
            'ëŠ¦ì€ë°© í•©ê³„': total_stats['late'].get(person, 0),
            'ë‹¹ì§ í•©ê³„': total_stats['duty'].get(person, 0)
        }
        for slot in slot_columns:
            row[f'{slot} í•©ê³„'] = total_stats['slots'][slot].get(person, 0)
        stats_data.append(row)

    stats_df = pd.DataFrame(stats_data)
    stats_df = stats_df.sort_values(by='ì¸ì›').reset_index(drop=True)

    # ì •ë³´ ì¶œë ¥
    person_info = {}
    max_assignments = {
        'ì´ë¥¸ë°© í•©ê³„': MAX_EARLY, 'ëŠ¦ì€ë°© í•©ê³„': MAX_LATE, 'ë‹¹ì§ í•©ê³„': MAX_DUTY,
        '8:30(1) í•©ê³„': MAX_ROOM, '8:30(2) í•©ê³„': MAX_ROOM, '8:30(4) í•©ê³„': MAX_ROOM, '8:30(7) í•©ê³„': MAX_ROOM,
        '9:00(10) í•©ê³„': MAX_ROOM, '9:00(11) í•©ê³„': MAX_ROOM, '9:00(12) í•©ê³„': MAX_ROOM,
        '9:30(8) í•©ê³„': MAX_ROOM, '9:30(5) í•©ê³„': MAX_ROOM, '9:30(6) í•©ê³„': MAX_ROOM,
        '10:00(9) í•©ê³„': MAX_ROOM, '10:00(3) í•©ê³„': MAX_ROOM,
        '13:30(3) í•©ê³„': MAX_ROOM, '13:30(4) í•©ê³„': MAX_ROOM, '13:30(9) í•©ê³„': MAX_ROOM, '13:30(2) í•©ê³„': MAX_ROOM
    }

    for slot in total_fixed_stats:
        for person, count in total_fixed_stats[slot].items():
            if count > 0:
                if person not in person_info:
                    person_info[person] = {'fixed': {}, 'priority': {}, 'sums': {}}
                person_info[person]['fixed'][slot] = count

    for slot in total_memo_stats:
        for person, count in total_memo_stats[slot].items():
            if count > 0:
                if person not in person_info:
                    person_info[person] = {'fixed': {}, 'priority': {}, 'sums': {}}
                person_info[person]['priority'][slot] = count

    for idx, row in stats_df.iterrows():
        person = row['ì¸ì›']
        if person not in person_info:
            person_info[person] = {'fixed': {}, 'priority': {}, 'sums': {}}
        for col in stats_df.columns[1:]:
            person_info[person]['sums'][col] = row[col]

    st.divider()
    st.write("### ğŸ‘¥ ì¸ì›ë³„ ìš°ì„ (ê³ ì •)ë°°ì • ì •ë³´")

    html_content = ""
    sorted_names = sorted(person_info.keys())

    merged_info = defaultdict(lambda: {"fixed": [], "priority": []})

    for person, info in person_info.items():
        base_name = re.sub(r'_\d+$', '', person)
        for slot, count in info['fixed'].items():
            merged_info[base_name]["fixed"].append(f"{slot} {count}ë²ˆ ê³ ì • ë°°ì¹˜")
        for slot, count in info['priority'].items():
            merged_info[base_name]["priority"].append(f"{slot} {count}ë²ˆ ìš°ì„ ë°°ì¹˜")

    html_content = ""
    sorted_names = sorted(merged_info.keys())

    for person in sorted_names:
        info = merged_info[person]
        output = [f"<span class='person'>{person}: </span>"]
        fixed_str = " / ".join(info["fixed"])
        priority_str = " / ".join(info["priority"])
        if fixed_str or priority_str:
            if fixed_str:
                output.append(fixed_str)
            if priority_str:
                output.append(f" / {priority_str}" if fixed_str else priority_str)
            html_content += f"<p>{''.join(output)}</p>"

    st.markdown(
        f"""
        <style>
        .custom-callout {{
            background-color: #f0f8ff;
            padding: 8px;
            border-radius: 6px;
            border-left: 4px solid #4682b4;
            box-shadow: 0px 2px 4px rgba(0, 0, 0, 0.1);
            margin-bottom: 4px;
            font-size: 14px;
            color: #2C3E50;
            line-height: 1.3;
        }}
        .custom-callout p {{
            margin: 0;
            padding: 2px 0;
            text-align: left;
        }}
        .person {{
            font-weight: bold;
            color: #2C3E50;
        }}
        </style>
        <div class="custom-callout">{html_content}</div>
        """,
        unsafe_allow_html=True
    )

    st.divider()
    st.write("### âš ï¸ ìµœëŒ€ ë°°ì • í•œê³„ ì´ˆê³¼ ê²½ê³ ")

    warnings = []
    for person in sorted_names:
        info = person_info[person]
        for slot_sum, count in info['sums'].items():
            max_count = max_assignments.get(slot_sum, float('inf'))
            if count > max_count:
                warnings.append(f"<span class='person'>{person}: </span>{slot_sum} = {count} (ìµœëŒ€ {max_count}ë²ˆ ì´ˆê³¼)")

    if warnings:
        warning_text = "".join([f"<p>{w}</p>" for w in warnings])
        html_content = f"""
        <div class="custom-callout warning-callout">
            {warning_text}
        </div>
        """
    else:
        html_content = """
        <div class="custom-callout warning-callout">
            <p>ëª¨ë“  ë°°ì •ì´ ì ì ˆí•œ í•œê³„ ë‚´ì— ìˆìŠµë‹ˆë‹¤.</p>
        </div>
        """

    st.markdown(
        f"""
        <style>
        .custom-callout {{
            padding: 8px;
            border-radius: 6px;
            box-shadow: 0px 2px 4px rgba(0, 0, 0, 0.1);
            margin-bottom: 4px;
            font-size: 14px;
            color: #2C3E50;
            line-height: 1.3;
        }}
        .custom-callout p {{
            margin: 0;
            padding: 2px 0;
            text-align: left;
        }}
        .person {{
            font-weight: bold;
            color: #2C3E50;
        }}
        .warning-callout {{
            background-color: #fff3cd;
            border-left: 4px solid #ffa500;
        }}
        </style>
        {html_content}
        """,
        unsafe_allow_html=True
    )

    st.divider()
    st.write("### í†µí•© ë°°ì¹˜ ê²°ê³¼")
    st.dataframe(result_df)

    # "ì¬ëœë¤í™”" ë²„íŠ¼
    if st.button("ì¬ëœë¤í™”"):
        st.session_state.clear()
        st.session_state.last_file_hash = file_hash
        st.rerun()

    st.divider()
    st.write("### ì¸ì›ë³„ ì „ì²´ í†µê³„")
    st.dataframe(stats_df)

    # ì—‘ì…€ ì›Œí¬ë¶ ìƒì„±
    output_wb = openpyxl.Workbook()
    schedule_sheet = output_wb.active
    schedule_sheet.title = "Schedule"

    default_font = Font(name="ë§‘ì€ ê³ ë”•", size=9)
    bold_font = Font(name="ë§‘ì€ ê³ ë”•", size=9, bold=True)
    magenta_bold_font = Font(name="ë§‘ì€ ê³ ë”•", size=9, bold=True, color="FF00FF")
    alignment_center = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    date_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    empty_row_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    weekday_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    saturday_with_person_fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    schedule_header_colors = {
        '8:30(1)_ë‹¹ì§': PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
        '8:30(2)': PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
        '8:30(4)': PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
        '8:30(7)': PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
        '9:00(10)': PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
        '9:00(11)': PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
        '9:00(12)': PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
        '9:30(8)': PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"),
        '9:30(5)': PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"),
        '9:30(6)': PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"),
        '10:00(9)': PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"),
        '10:00(3)': PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"),
        '13:30(3)_ë‹¹ì§': PatternFill(start_color="CC99FF", end_color="CC99FF", fill_type="solid"),
        '13:30(4)': PatternFill(start_color="CC99FF", end_color="CC99FF", fill_type="solid"),
        '13:30(9)': PatternFill(start_color="CC99FF", end_color="CC99FF", fill_type="solid"),
        '13:30(2)': PatternFill(start_color="CC99FF", end_color="CC99FF", fill_type="solid"),
        'ì˜¨ì½œ': PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    }

    date_header_cell = schedule_sheet['A1']
    date_header_cell.value = 'ë‚ ì§œ'
    date_header_cell.font = bold_font
    date_header_cell.alignment = alignment_center
    date_header_cell.border = border

    day_header_cell = schedule_sheet['B1']
    day_header_cell.value = 'ìš”ì¼'
    day_header_cell.font = bold_font
    day_header_cell.alignment = alignment_center
    date_header_cell.border = border

    for i, slot in enumerate(time_slots.keys(), 2):
        cell = schedule_sheet.cell(row=1, column=i+1, value=slot)
        cell.fill = schedule_header_colors.get(slot, PatternFill())
        cell.font = bold_font
        cell.alignment = alignment_center
        cell.border = border

    schedule_sheet.column_dimensions['A'].width = 12
    schedule_sheet.column_dimensions['B'].width = 8
    for col in range(3, len(time_slots.keys()) + 3):
        schedule_sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 10

    for i, row in enumerate(result_data, 2):
        date = row[0]
        date_obj = datetime.strptime(date, '%Y-%m-%d')
        formatted_date = date_obj.strftime('%mì›” %dì¼')
        
        has_person = any(x is not None and x != '' for x in row[2:])

        date_cell = schedule_sheet.cell(row=i, column=1, value=formatted_date)
        date_cell.fill = date_fill
        date_cell.font = bold_font
        date_cell.alignment = alignment_center
        date_cell.border = border

        day_of_week = row[1]
        day_cell = schedule_sheet.cell(row=i, column=2, value=day_of_week)
        if not has_person:
            day_cell.fill = empty_row_fill
        elif day_of_week == 'í† ìš”ì¼':
            day_cell.fill = saturday_with_person_fill
        elif day_of_week in ['ì›”ìš”ì¼', 'í™”ìš”ì¼', 'ìˆ˜ìš”ì¼', 'ëª©ìš”ì¼']:
            day_cell.fill = weekday_fill
        day_cell.font = default_font
        day_cell.alignment = alignment_center
        day_cell.border = border

        for j, value in enumerate(row[2:], 2):
            cell = schedule_sheet.cell(row=i, column=j+1, value=value)
            slot = list(time_slots.keys())[j-2]
            if slot in ['8:30(1)_ë‹¹ì§', '13:30(3)_ë‹¹ì§', 'ì˜¨ì½œ']:
                cell.font = magenta_bold_font
            else:
                cell.font = default_font
            if not has_person:
                cell.fill = empty_row_fill
            cell.alignment = alignment_center
            cell.border = border
            memo_key = (value, slot) if value else None
            if value and date in memo_mapping and memo_key in memo_mapping[date]:
                memo = memo_mapping[date][memo_key]
                cell.comment = Comment(memo, "Memo")

    stats_sheet = output_wb.create_sheet(title="Personnel_Stats")

    personnel_fill = PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid")
    total_sum_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    early_sum_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    late_sum_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    duty_sum_fill = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")
    slot_830_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    slot_900_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    slot_930_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    slot_1000_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    slot_1330_fill = PatternFill(start_color="CC99FF", end_color="CC99FF", fill_type="solid")

    
    headers = [
        'ì¸ì›', 'ì „ì²´ í•©ê³„', 'ì´ë¥¸ë°© í•©ê³„', 'ëŠ¦ì€ë°© í•©ê³„', 'ë‹¹ì§ í•©ê³„',
        '8:30(1) í•©ê³„', '8:30(2) í•©ê³„', '8:30(4) í•©ê³„', '8:30(7) í•©ê³„',
        '9:00(10) í•©ê³„', '9:00(11) í•©ê³„', '9:00(12) í•©ê³„',
        '9:30(8) í•©ê³„', '9:30(5) í•©ê³„', '9:30(6) í•©ê³„',
        '10:00(9) í•©ê³„', '10:00(3) í•©ê³„',
        '13:30(3) í•©ê³„', '13:30(4) í•©ê³„', '13:30(9) í•©ê³„', '13:30(2) í•©ê³„'
    ]

    for col, header in enumerate(headers, 1):
        cell = stats_sheet.cell(row=1, column=col, value=header)
        cell.font = bold_font
        cell.alignment = alignment_center
        cell.border = border
        if header == 'ì¸ì›':
            cell.fill = personnel_fill
        elif header == 'ì „ì²´ í•©ê³„':
            cell.fill = total_sum_fill
        elif header == 'ì´ë¥¸ë°© í•©ê³„':
            cell.fill = early_sum_fill
        elif header == 'ëŠ¦ì€ë°© í•©ê³„':
            cell.fill = late_sum_fill
        elif header == 'ë‹¹ì§ í•©ê³„':
            cell.fill = duty_sum_fill
        elif header in ['8:30(1) í•©ê³„', '8:30(2) í•©ê³„', '8:30(4) í•©ê³„', '8:30(7) í•©ê³„']:
            cell.fill = slot_830_fill
        elif header in ['9:00(10) í•©ê³„', '9:00(11) í•©ê³„', '9:00(12) í•©ê³„']:
            cell.fill = slot_900_fill
        elif header in ['9:30(8) í•©ê³„', '9:30(5) í•©ê³„', '9:30(6) í•©ê³„']:
            cell.fill = slot_930_fill
        elif header in ['10:00(9) í•©ê³„', '10:00(3) í•©ê³„']:
            cell.fill = slot_1000_fill
        elif header in ['13:30(3) í•©ê³„', '13:30(4) í•©ê³„', '13:30(9) í•©ê³„', '13:30(2) í•©ê³„']:
            cell.fill = slot_1330_fill

    for row_idx, row in enumerate(stats_df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = stats_sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = default_font
            cell.alignment = alignment_center
            cell.border = border
            header = headers[col_idx - 1]
            if header == 'ì¸ì›':
                cell.font = bold_font
                cell.fill = personnel_fill

    stats_sheet.column_dimensions['A'].width = 8
    for col in range(2, len(headers) + 1):
        stats_sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 10

    output_stream = BytesIO()
    output_wb.save(output_stream)
    output_stream.seek(0)

    today = datetime.today().strftime("%Y-%m-%d")
    st.divider()
    st.write("### ê²°ê³¼ ë‹¤ìš´ë¡œë“œ")
    st.write("- í†µí•© ë°°ì¹˜ ê²°ê³¼, ì¸ì›ë³„ ì „ì²´ í†µê³„ ì—‘ì…€ íŒŒì¼ì„ ë‹¤ìš´ë¡œë“œí•©ë‹ˆë‹¤.")
    st.download_button(
        label="ë‹¤ìš´ë¡œë“œ",
        data=output_stream,
        file_name=f"{today}_ë‚´ì‹œê²½ì‹¤ë°°ì •.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
